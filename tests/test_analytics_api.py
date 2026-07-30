"""Analytics API tests."""

from __future__ import annotations

import asyncio
import uuid
from collections.abc import AsyncGenerator, Generator
from datetime import UTC, datetime, timedelta, timezone
from typing import cast

import pytest
from fastapi.testclient import TestClient
from sqlalchemy import select, text
from sqlalchemy.engine import Connection
from sqlalchemy.ext.asyncio import AsyncSession, async_sessionmaker, create_async_engine
from sqlalchemy.pool import StaticPool
from sqlalchemy.sql.schema import Table

from app.core.security import create_token
from app.db.session import get_session
from app.main import app
from app.models.channel import Channel
from app.models.conversation import Conversation, Customer, CustomerIdentity, Message
from app.models.knowledge import KbCandidate, KbChunk, KbDocument
from app.models.ops import AIUsageEvent, Plan, Subscription, UsageCounter
from app.models.tenant import Tenant
from app.models.user import User

TENANT_ID = uuid.UUID("66666666-6666-4666-8666-666666666601")
OTHER_TENANT_ID = uuid.UUID("66666666-6666-4666-8666-666666666602")
USER_ID = uuid.UUID("66666666-6666-4666-8666-666666666603")
CHANNEL_ID = uuid.UUID("66666666-6666-4666-8666-666666666604")
CUSTOMER_ID = uuid.UUID("66666666-6666-4666-8666-666666666605")
AUTO_CUSTOMER_ID = uuid.UUID("66666666-6666-4666-8666-666666666606")
ESCALATED_CUSTOMER_ID = uuid.UUID("66666666-6666-4666-8666-666666666607")
OTHER_CHANNEL_ID = uuid.UUID("66666666-6666-4666-8666-666666666608")
OTHER_CUSTOMER_ID = uuid.UUID("66666666-6666-4666-8666-666666666609")


def create_table(sync_connection: Connection, table: object) -> None:
    cast(Table, table).create(sync_connection)


@pytest.fixture()
async def session_factory() -> AsyncGenerator[async_sessionmaker[AsyncSession], None]:
    engine = create_async_engine(
        "sqlite+aiosqlite://",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )
    async with engine.begin() as conn:
        for table in (
            Tenant.__table__,
            User.__table__,
            Channel.__table__,
            Customer.__table__,
            CustomerIdentity.__table__,
            Conversation.__table__,
            Message.__table__,
            AIUsageEvent.__table__,
            Plan.__table__,
            Subscription.__table__,
            UsageCounter.__table__,
            KbDocument.__table__,
            KbChunk.__table__,
            KbCandidate.__table__,
        ):
            await conn.run_sync(create_table, table)

    factory = async_sessionmaker(engine, expire_on_commit=False)
    try:
        yield factory
    finally:
        await engine.dispose()


@pytest.fixture()
def client(
    session_factory: async_sessionmaker[AsyncSession],
) -> Generator[TestClient, None, None]:
    async def override_get_session() -> AsyncGenerator[AsyncSession, None]:
        async with session_factory() as session:
            yield session

    app.dependency_overrides[get_session] = override_get_session
    try:
        yield TestClient(app)
    finally:
        app.dependency_overrides.pop(get_session, None)


def auth_headers() -> dict[str, str]:
    token = create_token(USER_ID, tenant_id=TENANT_ID, role="owner")
    return {"Authorization": f"Bearer {token}"}


async def seed_analytics_data(session_factory: async_sessionmaker[AsyncSession]) -> None:
    async with session_factory() as session:
        now = datetime.now(UTC)
        session.add_all(
            [
                Tenant(id=TENANT_ID, name="Demo", slug="demo", status="active"),
                Tenant(id=OTHER_TENANT_ID, name="Other", slug="other", status="active"),
                User(
                    id=USER_ID,
                    tenant_id=TENANT_ID,
                    email="owner@example.com",
                    full_name="Owner",
                    role="owner",
                    password_hash="test-password-hash",
                    status="active",
                ),
                Channel(
                    id=CHANNEL_ID,
                    tenant_id=TENANT_ID,
                    type="telegram",
                    name="Telegram",
                    status="active",
                    credentials_encrypted="fernet:test",
                    settings={},
                ),
                Channel(
                    id=OTHER_CHANNEL_ID,
                    tenant_id=OTHER_TENANT_ID,
                    type="telegram",
                    name="Other Telegram",
                    status="active",
                    credentials_encrypted="fernet:other",
                    settings={},
                ),
                Customer(
                    id=CUSTOMER_ID,
                    tenant_id=TENANT_ID,
                    display_name="Alina",
                    note="",
                ),
                Customer(
                    id=AUTO_CUSTOMER_ID,
                    tenant_id=TENANT_ID,
                    display_name="Auto customer",
                    note="",
                ),
                Customer(
                    id=ESCALATED_CUSTOMER_ID,
                    tenant_id=TENANT_ID,
                    display_name="Escalated customer",
                    note="",
                ),
                Customer(
                    id=OTHER_CUSTOMER_ID,
                    tenant_id=OTHER_TENANT_ID,
                    display_name="Other customer",
                    note="",
                ),
            ]
        )
        plan = Plan(
            code="demo",
            name="Demo",
            price_month=0,
            dialog_limit=500,
            channel_limit=1,
            features={},
        )
        session.add(plan)
        await session.flush()
        session.add(Subscription(tenant_id=TENANT_ID, plan_id=plan.id, status="trial"))
        session.add(
            UsageCounter(
                tenant_id=TENANT_ID,
                period=now.strftime("%Y-%m"),
                dialogs_count=9,
                ai_replies_count=2,
            )
        )

        open_conversation = Conversation(
            tenant_id=TENANT_ID,
            customer_id=CUSTOMER_ID,
            channel_id=CHANNEL_ID,
            status="open",
            last_message_at=now,
            last_message_preview="Open",
            unread_count=1,
        )
        auto_conversation = Conversation(
            tenant_id=TENANT_ID,
            customer_id=AUTO_CUSTOMER_ID,
            channel_id=CHANNEL_ID,
            status="auto",
            last_message_at=now,
            last_message_preview="Auto",
            unread_count=0,
        )
        escalated_conversation = Conversation(
            tenant_id=TENANT_ID,
            customer_id=ESCALATED_CUSTOMER_ID,
            channel_id=CHANNEL_ID,
            status="escalated",
            last_message_at=now,
            last_message_preview="Escalated",
            unread_count=2,
        )
        other_conversation = Conversation(
            tenant_id=OTHER_TENANT_ID,
            customer_id=OTHER_CUSTOMER_ID,
            channel_id=OTHER_CHANNEL_ID,
            status="auto",
            last_message_at=now,
            last_message_preview="Other",
            unread_count=0,
        )
        session.add_all(
            [
                open_conversation,
                auto_conversation,
                escalated_conversation,
                other_conversation,
            ]
        )
        await session.flush()

        session.add_all(
            [
                Message(
                    tenant_id=TENANT_ID,
                    conversation_id=auto_conversation.id,
                    direction="inbound",
                    sender_type="customer",
                    text="Question 1",
                    attachments={},
                    status="received",
                    confidence=None,
                    ai_meta={},
                    created_at=now - timedelta(seconds=70),
                ),
                Message(
                    tenant_id=TENANT_ID,
                    conversation_id=auto_conversation.id,
                    direction="outbound",
                    sender_type="ai",
                    text="Answer 1",
                    attachments={},
                    status="sent",
                    confidence=0.8,
                    ai_meta={"provider": "mock"},
                    created_at=now - timedelta(seconds=10),
                ),
                Message(
                    tenant_id=TENANT_ID,
                    conversation_id=escalated_conversation.id,
                    direction="inbound",
                    sender_type="customer",
                    text="Question 2",
                    attachments={},
                    status="received",
                    confidence=None,
                    ai_meta={},
                    created_at=now - timedelta(seconds=120),
                ),
                Message(
                    tenant_id=TENANT_ID,
                    conversation_id=escalated_conversation.id,
                    direction="outbound",
                    sender_type="manager",
                    text="Answer 2",
                    attachments={},
                    status="sent",
                    confidence=None,
                    ai_meta={},
                    created_at=now - timedelta(seconds=30),
                ),
                Message(
                    tenant_id=OTHER_TENANT_ID,
                    conversation_id=other_conversation.id,
                    direction="outbound",
                    sender_type="ai",
                    text="Other answer",
                    attachments={},
                    status="sent",
                    confidence=1.0,
                    ai_meta={},
                    created_at=now,
                ),
            ]
        )

        ready_document = KbDocument(
            tenant_id=TENANT_ID,
            title="FAQ",
            source_type="manual",
            status="ready",
            version=1,
        )
        archived_document = KbDocument(
            tenant_id=TENANT_ID,
            title="Old",
            source_type="manual",
            status="archived",
            version=1,
        )
        session.add_all([ready_document, archived_document])
        await session.flush()
        session.add_all(
            [
                KbChunk(
                    tenant_id=TENANT_ID,
                    document_id=ready_document.id,
                    text="Chunk 1",
                    position=0,
                    token_count=2,
                    tags={},
                    version=1,
                ),
                KbChunk(
                    tenant_id=TENANT_ID,
                    document_id=ready_document.id,
                    text="Chunk 2",
                    position=1,
                    token_count=2,
                    tags={},
                    version=1,
                ),
                KbCandidate(
                    tenant_id=TENANT_ID,
                    conversation_id=escalated_conversation.id,
                    question="Q",
                    answer="A",
                    suggested_by="manager",
                    status="pending",
                ),
            ]
        )

        await session.commit()


def test_analytics_overview_returns_tenant_kpis(
    client: TestClient,
    session_factory: async_sessionmaker[AsyncSession],
) -> None:
    asyncio.run(seed_analytics_data(session_factory))

    response = client.get("/api/v1/analytics/overview", headers=auth_headers())

    assert response.status_code == 200, response.text
    data = response.json()
    assert data["dialogs_total"] == 2
    assert data["dialogs_open"] == 0
    assert data["dialogs_auto"] == 1
    assert data["dialogs_escalated"] == 1
    assert data["auto_reply_rate"] == 0.5
    assert data["escalation_rate"] == 0.5
    assert data["avg_response_sec"] == 75
    assert data["avg_ai_confidence"] == 0.8
    assert data["ai_replies_count"] == 1
    assert data["manager_replies_count"] == 1
    assert data["inbound_messages_count"] == 2
    assert data["dialogs_used"] == 9
    assert data["dialogs_limit"] == 500
    assert data["knowledge_documents_ready"] == 1
    assert data["knowledge_chunks_count"] == 2
    assert data["pending_candidates_count"] == 1
    assert data["status_breakdown"] == [
        {"status": "auto", "count": 1},
        {"status": "escalated", "count": 1},
    ]


def test_detailed_analytics_export_contains_customer_conversation_and_message_sheets(
    client: TestClient,
    session_factory: async_sessionmaker[AsyncSession],
) -> None:
    from io import BytesIO

    from openpyxl import load_workbook

    asyncio.run(seed_analytics_data(session_factory))

    response = client.get("/api/v1/analytics/export", headers=auth_headers())

    assert response.status_code == 200, response.text
    assert response.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    workbook = load_workbook(BytesIO(response.content), data_only=True)
    assert workbook.sheetnames == [
        "Сводка",
        "По дням",
        "Клиенты",
        "Диалоги",
        "Сообщения",
        "Использование AI",
    ]
    customer_names = {
        workbook["Клиенты"].cell(row=row, column=2).value
        for row in range(2, workbook["Клиенты"].max_row + 1)
    }
    assert customer_names == {"Auto customer", "Escalated customer"}
    assert workbook["Сообщения"].max_row == 5
    assert workbook["Сводка"]["B19"].number_format == '#,##0.00 "₽"'
    assert workbook["Сводка"]["B3"].number_format == "yyyy-mm-dd h:mm:ss"


def test_detailed_analytics_export_works_for_empty_period(
    client: TestClient,
    session_factory: async_sessionmaker[AsyncSession],
) -> None:
    from io import BytesIO

    from openpyxl import load_workbook

    async def seed_empty_tenant() -> None:
        async with session_factory() as session:
            session.add(Tenant(id=TENANT_ID, name="Empty", slug="empty", status="active"))
            session.add(
                User(
                    id=USER_ID,
                    tenant_id=TENANT_ID,
                    email="owner@example.com",
                    full_name="Owner",
                    role="owner",
                    password_hash="test-password-hash",
                    status="active",
                )
            )
            await session.commit()

    asyncio.run(seed_empty_tenant())

    response = client.get(
        "/api/v1/analytics/export?from=2026-07-01&to=2026-07-01",
        headers=auth_headers(),
    )

    assert response.status_code == 200, response.text
    assert response.headers["content-disposition"] == (
        'attachment; filename="autopilot-analytics-2026-07-01-2026-07-01.xlsx"'
    )
    workbook = load_workbook(BytesIO(response.content), data_only=True)
    assert workbook["Сводка"]["B4"].value == 0
    assert workbook["Сводка"]["B9"].value == 0
    assert workbook["По дням"].max_row == 3
    assert workbook["По дням"]["A3"].value == "ИТОГО ЗА ПЕРИОД"
    assert workbook["Клиенты"].max_row == 1
    assert workbook["Диалоги"].max_row == 1
    assert workbook["Сообщения"].max_row == 1
    assert workbook["Использование AI"].max_row == 1


def test_detailed_analytics_export_stays_available_before_usage_migration(
    client: TestClient,
    session_factory: async_sessionmaker[AsyncSession],
) -> None:
    from io import BytesIO

    from openpyxl import load_workbook

    asyncio.run(seed_analytics_data(session_factory))

    async def remove_usage_ledger() -> None:
        async with session_factory() as session:
            await session.execute(text("DROP TABLE ai_usage_event"))
            await session.commit()

    asyncio.run(remove_usage_ledger())

    response = client.get("/api/v1/analytics/export", headers=auth_headers())

    assert response.status_code == 200, response.text
    workbook = load_workbook(BytesIO(response.content), data_only=True)
    assert workbook["Клиенты"].max_row == 3
    assert workbook["Сообщения"].max_row == 5
    assert workbook["Использование AI"].max_row == 1


def test_detailed_analytics_export_works_for_one_dialog_without_usage(
    client: TestClient,
    session_factory: async_sessionmaker[AsyncSession],
) -> None:
    from io import BytesIO

    from openpyxl import load_workbook

    async def seed_one_dialog() -> None:
        async with session_factory() as session:
            session.add(Tenant(id=TENANT_ID, name="Single", slug="single", status="active"))
            session.add(
                User(
                    id=USER_ID,
                    tenant_id=TENANT_ID,
                    email="owner@example.com",
                    full_name="Owner",
                    role="owner",
                    password_hash="test-password-hash",
                    status="active",
                )
            )
            session.add(
                Channel(
                    id=CHANNEL_ID,
                    tenant_id=TENANT_ID,
                    type="telegram",
                    name="Telegram",
                    status="active",
                    credentials_encrypted="fernet:test",
                    settings={},
                )
            )
            session.add(
                Customer(
                    id=CUSTOMER_ID,
                    tenant_id=TENANT_ID,
                    display_name="Один клиент",
                    note="",
                )
            )
            await session.flush()
            conversation = Conversation(
                tenant_id=TENANT_ID,
                customer_id=CUSTOMER_ID,
                channel_id=CHANNEL_ID,
                status="answered",
                last_message_at=datetime(2026, 7, 15, 9, 2, tzinfo=UTC),
                last_message_preview="Ответ",
                unread_count=0,
            )
            session.add(conversation)
            await session.flush()
            session.add_all(
                [
                    Message(
                        tenant_id=TENANT_ID,
                        conversation_id=conversation.id,
                        direction="inbound",
                        sender_type="customer",
                        text="Вопрос",
                        attachments={},
                        status="received",
                        confidence=None,
                        ai_meta={},
                        created_at=datetime(2026, 7, 15, 9, 0, tzinfo=UTC),
                    ),
                    Message(
                        tenant_id=TENANT_ID,
                        conversation_id=conversation.id,
                        direction="outbound",
                        sender_type="manager",
                        text="Ответ",
                        attachments={},
                        status="sent",
                        confidence=None,
                        ai_meta={},
                        created_at=datetime(2026, 7, 15, 9, 2, tzinfo=UTC),
                    ),
                ]
            )
            await session.commit()

    asyncio.run(seed_one_dialog())

    response = client.get(
        "/api/v1/analytics/export?from=2026-07-15&to=2026-07-15",
        headers=auth_headers(),
    )

    assert response.status_code == 200, response.text
    workbook = load_workbook(BytesIO(response.content), data_only=True)
    assert workbook["Клиенты"].max_row == 2
    assert workbook["Клиенты"]["B2"].value == "Один клиент"
    assert workbook["Клиенты"]["L2"].value == "Нет AI-ответов"
    assert workbook["Клиенты"]["R2"].value == 0
    assert workbook["Диалоги"].max_row == 2
    assert workbook["Сообщения"].max_row == 3
    assert workbook["Сообщения"]["T2"].value is None
    assert workbook["Сообщения"]["T3"].value is None
    assert workbook["Использование AI"].max_row == 1


def test_detailed_analytics_export_marks_missing_historical_token_data(
    client: TestClient,
    session_factory: async_sessionmaker[AsyncSession],
) -> None:
    from io import BytesIO

    from openpyxl import load_workbook

    asyncio.run(seed_analytics_data(session_factory))

    response = client.get("/api/v1/analytics/export", headers=auth_headers())

    assert response.status_code == 200, response.text
    workbook = load_workbook(BytesIO(response.content), data_only=True)
    summary = workbook["Сводка"]
    assert summary["B11"].value == 0
    assert summary["B12"].value == "Нет исторических данных"
    assert summary["B13"].value is None
    daily = workbook["По дням"]
    active_rows = [row for row in range(2, daily.max_row) if daily.cell(row, 5).value]
    assert active_rows
    active_row = active_rows[0]
    assert daily.cell(active_row, 9).value == "Нет исторических данных"
    assert daily.cell(active_row, 10).value is None
    assert daily.cell(active_row, 16).value is None
    assert daily.cell(daily.max_row, 1).value == "ИТОГО ЗА ПЕРИОД"
    assert daily.cell(daily.max_row, 9).value == "Нет исторических данных"
    assert daily.auto_filter.ref.endswith(str(daily.max_row - 1))
    assert daily.freeze_panes == "B2"
    assert daily["A1"].alignment.wrap_text is True
    assert daily.row_dimensions[1].height == 42


def test_detailed_analytics_export_does_not_treat_empty_usage_as_measured(
    client: TestClient,
    session_factory: async_sessionmaker[AsyncSession],
) -> None:
    from io import BytesIO

    from openpyxl import load_workbook

    asyncio.run(seed_analytics_data(session_factory))

    async def add_empty_usage_object() -> None:
        async with session_factory() as session:
            result = await session.execute(
                select(Message).where(
                    Message.tenant_id == TENANT_ID,
                    Message.sender_type == "ai",
                )
            )
            message = result.scalars().first()
            assert message is not None
            message.ai_meta = {**(message.ai_meta or {}), "usage": {}}
            await session.commit()

    asyncio.run(add_empty_usage_object())
    response = client.get("/api/v1/analytics/export", headers=auth_headers())

    assert response.status_code == 200, response.text
    workbook = load_workbook(BytesIO(response.content), data_only=True)
    assert workbook["Сводка"]["B11"].value == 0
    assert workbook["Сводка"]["B12"].value == "Нет исторических данных"
    assert workbook["Сводка"]["B13"].value is None


def test_detailed_analytics_export_can_omit_message_sheet(
    client: TestClient,
    session_factory: async_sessionmaker[AsyncSession],
) -> None:
    asyncio.run(seed_analytics_data(session_factory))

    response = client.get(
        "/api/v1/analytics/export?include_messages=false",
        headers=auth_headers(),
    )

    assert response.status_code == 200, response.text
    from io import BytesIO

    from openpyxl import load_workbook

    workbook = load_workbook(BytesIO(response.content), read_only=True)
    assert "Сообщения" not in workbook.sheetnames


def test_detailed_analytics_export_rejects_invalid_period(
    client: TestClient,
    session_factory: async_sessionmaker[AsyncSession],
) -> None:
    asyncio.run(seed_analytics_data(session_factory))

    inverted = client.get(
        "/api/v1/analytics/export?from=2026-07-30&to=2026-07-01",
        headers=auth_headers(),
    )
    too_long = client.get(
        "/api/v1/analytics/export?from=2025-07-01&to=2026-07-30",
        headers=auth_headers(),
    )

    assert inverted.status_code == 422
    assert too_long.status_code == 422


def test_analytics_counts_answered_conversation_as_open(
    client: TestClient,
    session_factory: async_sessionmaker[AsyncSession],
) -> None:
    asyncio.run(seed_analytics_data(session_factory))

    async def mark_open_conversation_answered() -> None:
        async with session_factory() as session:
            result = await session.execute(
                select(Conversation).where(
                    Conversation.tenant_id == TENANT_ID,
                    Conversation.status == "open",
                )
            )
            conversation = result.scalar_one()
            conversation.status = "answered"
            session.add(
                Message(
                    tenant_id=TENANT_ID,
                    conversation_id=conversation.id,
                    direction="inbound",
                    sender_type="customer",
                    text="Новый вопрос",
                    attachments={},
                    status="received",
                    confidence=None,
                    ai_meta={},
                    created_at=datetime.now(UTC),
                )
            )
            await session.commit()

    asyncio.run(mark_open_conversation_answered())
    response = client.get("/api/v1/analytics/overview", headers=auth_headers())

    assert response.status_code == 200, response.text
    data = response.json()
    assert data["dialogs_open"] == 1
    assert {"status": "answered", "count": 1} in data["status_breakdown"]


def test_analytics_counts_unique_active_dialogs_by_moscow_day_and_fills_zeroes(
    client: TestClient,
    session_factory: async_sessionmaker[AsyncSession],
) -> None:
    asyncio.run(seed_analytics_data(session_factory))

    async def arrange_dates() -> None:
        async with session_factory() as session:
            conversations = list(
                (
                    await session.execute(
                        select(Conversation).where(Conversation.tenant_id == TENANT_ID)
                    )
                )
                .scalars()
                .all()
            )
            by_status = {conversation.status: conversation for conversation in conversations}
            by_status["open"].created_at = datetime(2026, 7, 1, 23, 59, tzinfo=UTC)
            by_status["auto"].created_at = datetime(2026, 7, 20, 0, 0, tzinfo=UTC)
            by_status["escalated"].created_at = datetime(2026, 7, 22, 0, 0, tzinfo=UTC)

            messages = list(
                (await session.execute(select(Message).where(Message.tenant_id == TENANT_ID)))
                .scalars()
                .all()
            )
            for message in messages:
                message.created_at = (
                    datetime(2026, 7, 20, 8, 0, tzinfo=UTC)
                    if message.conversation_id == by_status["auto"].id
                    else datetime(2026, 7, 22, 8, 0, tzinfo=UTC)
                )
            await session.commit()

    asyncio.run(arrange_dates())

    response = client.get(
        "/api/v1/analytics/overview?from=2026-07-20&to=2026-07-23",
        headers=auth_headers(),
    )

    assert response.status_code == 200, response.text
    data = response.json()
    assert data["date_from"] == "2026-07-20"
    assert data["date_to"] == "2026-07-23"
    assert data["dialogs_total"] == 2
    assert data["daily_series"] == [
        {"date": "2026-07-20", "dialogs": 1},
        {"date": "2026-07-21", "dialogs": 0},
        {"date": "2026-07-22", "dialogs": 1},
        {"date": "2026-07-23", "dialogs": 0},
    ]


def test_analytics_counts_reused_dialog_again_on_another_day_but_once_per_day(
    client: TestClient,
    session_factory: async_sessionmaker[AsyncSession],
) -> None:
    asyncio.run(seed_analytics_data(session_factory))

    async def arrange_repeated_inbound_messages() -> None:
        async with session_factory() as session:
            conversation = (
                await session.execute(
                    select(Conversation).where(
                        Conversation.tenant_id == TENANT_ID,
                        Conversation.status == "auto",
                    )
                )
            ).scalar_one()
            conversation.created_at = datetime(2026, 7, 1, tzinfo=UTC)
            existing_messages = list(
                (
                    await session.execute(
                        select(Message).where(Message.tenant_id == TENANT_ID)
                    )
                )
                .scalars()
                .all()
            )
            for message in existing_messages:
                message.created_at = datetime(2026, 7, 20, 12, 0, tzinfo=UTC)
            session.add_all(
                [
                    Message(
                        tenant_id=TENANT_ID,
                        conversation_id=conversation.id,
                        direction="inbound",
                        sender_type="customer",
                        text="Повторный вопрос 1",
                        attachments={},
                        status="received",
                        confidence=None,
                        ai_meta={},
                        created_at=datetime(2026, 7, 29, 22, 5, tzinfo=UTC),
                    ),
                    Message(
                        tenant_id=TENANT_ID,
                        conversation_id=conversation.id,
                        direction="inbound",
                        sender_type="customer",
                        text="Повторный вопрос 2",
                        attachments={},
                        status="received",
                        confidence=None,
                        ai_meta={},
                        created_at=datetime(2026, 7, 29, 22, 10, tzinfo=UTC),
                    ),
                    Message(
                        tenant_id=TENANT_ID,
                        conversation_id=conversation.id,
                        direction="inbound",
                        sender_type="customer",
                        text="Вопрос на следующий день",
                        attachments={},
                        status="received",
                        confidence=None,
                        ai_meta={},
                        created_at=datetime(2026, 7, 30, 21, 5, tzinfo=UTC),
                    ),
                ]
            )
            await session.commit()

    asyncio.run(arrange_repeated_inbound_messages())
    response = client.get(
        "/api/v1/analytics/overview?from=2026-07-30&to=2026-07-31",
        headers=auth_headers(),
    )

    assert response.status_code == 200, response.text
    data = response.json()
    assert data["dialogs_total"] == 1
    assert data["inbound_messages_count"] == 3
    assert data["daily_series"] == [
        {"date": "2026-07-30", "dialogs": 1},
        {"date": "2026-07-31", "dialogs": 1},
    ]


def test_analytics_auto_reply_rate_cannot_include_another_tenants_conversation(
    client: TestClient,
    session_factory: async_sessionmaker[AsyncSession],
) -> None:
    asyncio.run(seed_analytics_data(session_factory))

    async def add_cross_tenant_message() -> None:
        async with session_factory() as session:
            tenant_conversation = (
                await session.execute(
                    select(Conversation).where(
                        Conversation.tenant_id == TENANT_ID,
                        Conversation.status == "auto",
                    )
                )
            ).scalar_one()
            other_conversation = (
                await session.execute(
                    select(Conversation).where(Conversation.tenant_id == OTHER_TENANT_ID)
                )
            ).scalar_one()
            now = datetime.now(UTC)
            for conversation in (
                tenant_conversation,
                other_conversation,
            ):
                conversation.created_at = now
            session.add(
                Message(
                    tenant_id=TENANT_ID,
                    conversation_id=other_conversation.id,
                    direction="outbound",
                    sender_type="ai",
                    text="Must not affect tenant analytics",
                    attachments={},
                    status="sent",
                    confidence=1.0,
                    ai_meta={},
                    created_at=now,
                )
            )
            await session.commit()

    asyncio.run(add_cross_tenant_message())

    today = datetime.now(UTC).astimezone(timezone(timedelta(hours=3))).date().isoformat()
    response = client.get(
        f"/api/v1/analytics/overview?from={today}&to={today}",
        headers=auth_headers(),
    )

    assert response.status_code == 200, response.text
    data = response.json()
    assert data["auto_reply_rate"] == 0.5


def test_analytics_rejects_inverted_date_range(
    client: TestClient,
    session_factory: async_sessionmaker[AsyncSession],
) -> None:
    asyncio.run(seed_analytics_data(session_factory))

    response = client.get(
        "/api/v1/analytics/overview?from=2026-07-23&to=2026-07-20",
        headers=auth_headers(),
    )

    assert response.status_code == 422


def test_analytics_overview_returns_zeroes_for_empty_tenant(
    client: TestClient,
    session_factory: async_sessionmaker[AsyncSession],
) -> None:
    async def seed_empty_tenant() -> None:
        async with session_factory() as session:
            session.add(Tenant(id=TENANT_ID, name="Empty", slug="empty", status="active"))
            session.add(
                User(
                    id=USER_ID,
                    tenant_id=TENANT_ID,
                    email="owner@example.com",
                    full_name="Owner",
                    role="owner",
                    password_hash="test-password-hash",
                    status="active",
                )
            )
            await session.commit()

    asyncio.run(seed_empty_tenant())

    response = client.get("/api/v1/analytics/overview", headers=auth_headers())

    assert response.status_code == 200, response.text
    data = response.json()
    assert data["dialogs_total"] == 0
    assert data["auto_reply_rate"] == 0
    assert data["avg_response_sec"] == 0
    assert data["status_breakdown"] == []
