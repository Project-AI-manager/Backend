"""Analytics dashboard endpoints. Screen: /analytics."""

import io
from collections import Counter
from dataclasses import dataclass
from datetime import UTC, date, datetime, time, timedelta, timezone
from typing import Annotated, TypeGuard
from uuid import UUID

from fastapi import APIRouter, HTTPException, Query
from fastapi.responses import StreamingResponse
from openpyxl import Workbook  # type: ignore[import-untyped]
from openpyxl.styles import (  # type: ignore[import-untyped]
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter  # type: ignore[import-untyped]
from sqlalchemy import inspect, select
from sqlalchemy.exc import SQLAlchemyError

from app.api.deps import CurrentUser, SessionDep, tenant_id_from_user
from app.models.channel import Channel
from app.models.conversation import Conversation, Customer, CustomerIdentity, Message
from app.models.knowledge import KbCandidate, KbChunk, KbDocument
from app.models.ops import AIUsageEvent, Plan, Subscription, UsageCounter
from app.schemas.analytics import (
    AnalyticsDailySeriesItem,
    AnalyticsOverviewResponse,
    AnalyticsStatusBreakdownItem,
)

router = APIRouter()
DEFAULT_PERIOD_DAYS = 7
MAX_PERIOD_DAYS = 366
ANALYTICS_TIMEZONE = timezone(timedelta(hours=3), name="Europe/Moscow")


@dataclass(frozen=True)
class _TokenCoverage:
    recorded_replies: int
    label: str


@router.get("/export")
async def export_detailed_analytics(
    user: CurrentUser,
    session: SessionDep,
    from_date: Annotated[date | None, Query(alias="from")] = None,
    to_date: Annotated[date | None, Query(alias="to")] = None,
    include_messages: bool = True,
) -> StreamingResponse:
    tenant_id = tenant_id_from_user(user)
    date_from, date_to = _resolve_period(from_date, to_date)
    starts_at, ends_at = _analytics_bounds(date_from, date_to)
    messages = await _tenant_messages(session, tenant_id, starts_at, ends_at)
    conversation_ids = {message.conversation_id for message in messages}
    conversations = await _tenant_conversations(session, tenant_id, conversation_ids)
    customers = await _customers(session, tenant_id, {item.customer_id for item in conversations})
    channels = await _channels(session, tenant_id, {item.channel_id for item in conversations})
    identities = await _identities(session, {item.id for item in customers.values()})
    usage_events = await _usage_events(session, tenant_id, starts_at, ends_at)

    workbook = _build_analytics_workbook(
        date_from=date_from,
        date_to=date_to,
        conversations=conversations,
        messages=messages,
        customers=customers,
        channels=channels,
        identities=identities,
        usage_events=usage_events,
        include_messages=include_messages,
    )
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    filename = f"autopilot-analytics-{date_from.isoformat()}-{date_to.isoformat()}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/overview", response_model=AnalyticsOverviewResponse)
async def overview(
    user: CurrentUser,
    session: SessionDep,
    from_date: Annotated[date | None, Query(alias="from")] = None,
    to_date: Annotated[date | None, Query(alias="to")] = None,
) -> AnalyticsOverviewResponse:
    tenant_id = tenant_id_from_user(user)
    date_from, date_to = _resolve_period(from_date, to_date)
    starts_at, ends_at = _analytics_bounds(date_from, date_to)

    messages = await _tenant_messages(session, tenant_id, starts_at, ends_at)
    inbound_messages = [message for message in messages if message.direction == "inbound"]
    active_conversation_ids = {message.conversation_id for message in inbound_messages}
    conversations = await _tenant_conversations(
        session,
        tenant_id,
        active_conversation_ids,
    )
    dialogs_used, dialogs_limit = await _usage_and_limit(session, tenant_id)
    (
        knowledge_documents_ready,
        knowledge_chunks_count,
        pending_candidates_count,
    ) = await _knowledge_counts(session, tenant_id)

    status_counts = _status_counts(conversations)
    dialogs_total = len(conversations)
    ai_replies = [message for message in messages if message.sender_type == "ai"]
    manager_replies = [message for message in messages if message.sender_type == "manager"]
    conversation_ids = {conversation.id for conversation in conversations}
    conversations_with_ai = {
        message.conversation_id
        for message in ai_replies
        if message.conversation_id in conversation_ids
    }
    confidence_values = [
        message.confidence for message in ai_replies if isinstance(message.confidence, (int, float))
    ]

    return AnalyticsOverviewResponse(
        date_from=date_from,
        date_to=date_to,
        dialogs_total=dialogs_total,
        dialogs_open=status_counts.get("open", 0) + status_counts.get("answered", 0),
        dialogs_auto=status_counts.get("auto", 0),
        dialogs_escalated=status_counts.get("escalated", 0),
        dialogs_closed=status_counts.get("closed", 0),
        auto_reply_rate=_ratio(len(conversations_with_ai), dialogs_total),
        escalation_rate=_ratio(status_counts.get("escalated", 0), dialogs_total),
        avg_response_sec=_average_response_sec(messages),
        avg_ai_confidence=round(sum(confidence_values) / len(confidence_values), 4)
        if confidence_values
        else 0.0,
        ai_replies_count=len(ai_replies),
        manager_replies_count=len(manager_replies),
        inbound_messages_count=len(inbound_messages),
        dialogs_used=dialogs_used,
        dialogs_limit=dialogs_limit,
        knowledge_documents_ready=knowledge_documents_ready,
        knowledge_chunks_count=knowledge_chunks_count,
        pending_candidates_count=pending_candidates_count,
        status_breakdown=[
            AnalyticsStatusBreakdownItem(status=status, count=count)
            for status, count in sorted(status_counts.items())
        ],
        daily_series=_daily_series(inbound_messages, date_from, date_to),
    )


async def _tenant_conversations(
    session: SessionDep,
    tenant_id: UUID,
    conversation_ids: set[UUID],
) -> list[Conversation]:
    if not conversation_ids:
        return []
    result = await session.execute(
        select(Conversation).where(
            Conversation.tenant_id == tenant_id,
            Conversation.id.in_(conversation_ids),
        )
    )
    return list(result.scalars().all())


async def _tenant_messages(
    session: SessionDep,
    tenant_id: UUID,
    starts_at: datetime,
    ends_at: datetime,
) -> list[Message]:
    result = await session.execute(
        select(Message)
        .join(Conversation, Message.conversation_id == Conversation.id)
        .where(
            Message.tenant_id == tenant_id,
            Conversation.tenant_id == tenant_id,
            Message.created_at >= starts_at,
            Message.created_at < ends_at,
        )
        .order_by(Message.created_at)
    )
    return list(result.scalars().all())


async def _customers(
    session: SessionDep, tenant_id: UUID, customer_ids: set[UUID]
) -> dict[UUID, Customer]:
    if not customer_ids:
        return {}
    result = await session.execute(
        select(Customer).where(Customer.tenant_id == tenant_id, Customer.id.in_(customer_ids))
    )
    return {item.id: item for item in result.scalars().all()}


async def _channels(
    session: SessionDep, tenant_id: UUID, channel_ids: set[UUID]
) -> dict[UUID, Channel]:
    if not channel_ids:
        return {}
    result = await session.execute(
        select(Channel).where(Channel.tenant_id == tenant_id, Channel.id.in_(channel_ids))
    )
    return {item.id: item for item in result.scalars().all()}


async def _identities(session: SessionDep, customer_ids: set[UUID]) -> dict[UUID, str]:
    if not customer_ids:
        return {}
    result = await session.execute(
        select(CustomerIdentity).where(CustomerIdentity.customer_id.in_(customer_ids))
    )
    return {item.customer_id: item.external_user_id for item in result.scalars().all()}


async def _usage_events(
    session: SessionDep, tenant_id: UUID, starts_at: datetime, ends_at: datetime
) -> list[AIUsageEvent]:
    # Keep exports available during a rolling deployment where the application
    # code can start a little earlier than the optional usage-ledger migration.
    # Message-level analytics still has ai_meta as a best-effort fallback.
    connection = await session.connection()
    has_usage_table = await connection.run_sync(
        lambda sync_connection: inspect(sync_connection).has_table(AIUsageEvent.__tablename__)
    )
    if not has_usage_table:
        return []
    try:
        result = await session.execute(
            select(AIUsageEvent)
            .where(
                AIUsageEvent.tenant_id == tenant_id,
                AIUsageEvent.created_at >= starts_at,
                AIUsageEvent.created_at < ends_at,
            )
            .order_by(AIUsageEvent.created_at)
        )
    except SQLAlchemyError:
        await session.rollback()
        return []
    return list(result.scalars().all())


def _build_analytics_workbook(
    *,
    date_from: date,
    date_to: date,
    conversations: list[Conversation],
    messages: list[Message],
    customers: dict[UUID, Customer],
    channels: dict[UUID, Channel],
    identities: dict[UUID, str],
    usage_events: list[AIUsageEvent],
    include_messages: bool,
) -> Workbook:
    workbook = Workbook()
    summary = workbook.active
    summary.title = "Сводка"
    by_conversation = _group_messages(messages)
    usage_by_conversation = _group_usage(usage_events)
    active_customer_ids = {item.customer_id for item in conversations}
    period_ai_messages = [message for message in messages if message.sender_type == "ai"]
    period_coverage = _token_coverage(period_ai_messages, usage_events)
    period_token_values = _daily_token_values(period_ai_messages, usage_events)
    response_times = _response_times(messages)
    summary_rows: list[list] = [
        ["Отчёт Автопилота", f"{date_from.isoformat()} — {date_to.isoformat()}"],
        ["Сформировано", datetime.now(ANALYTICS_TIMEZONE).replace(tzinfo=None)],
        ["Уникальных клиентов", len(active_customer_ids)],
        ["Активных диалогов", len(conversations)],
        ["Сообщений клиентов", sum(item.direction == "inbound" for item in messages)],
        ["Ответов Автопилота", sum(item.sender_type == "ai" for item in messages)],
        ["Ответов менеджеров", sum(item.sender_type == "manager" for item in messages)],
        ["Всего сообщений", len(messages)],
        [
            "Среднее время ответа, сек",
            round(sum(response_times) / len(response_times)) if response_times else 0,
        ],
        ["AI-ответов с токенами", period_coverage.recorded_replies],
        ["Статус данных по токенам", period_coverage.label],
        ["Input tokens", period_token_values[0]],
        ["Cached input tokens", period_token_values[1]],
        ["Cache write tokens", period_token_values[2]],
        ["Reasoning tokens", period_token_values[3]],
        ["Output tokens (включая reasoning)", period_token_values[4]],
        ["Всего токенов", period_token_values[5]],
        ["Себестоимость API, ₽", period_token_values[6]],
        ["Списание клиенту, ₽", period_token_values[7]],
    ]
    _write_sheet(summary, ["Показатель", "Значение"], summary_rows)
    for row_index in (19, 20):
        summary.cell(row=row_index, column=2).number_format = '#,##0.00 "₽"'

    daily = workbook.create_sheet("По дням")
    daily_rows = []
    conversation_customer_ids = {
        conversation.id: conversation.customer_id for conversation in conversations
    }
    days_count = (date_to - date_from).days + 1
    for offset in range(days_count):
        current = date_from + timedelta(days=offset)
        day_messages = [m for m in messages if _as_analytics_date(m.created_at) == current]
        day_usage = [u for u in usage_events if _as_analytics_date(u.created_at) == current]
        day_ai_messages = [message for message in day_messages if message.sender_type == "ai"]
        token_coverage = _token_coverage(day_ai_messages, day_usage)
        token_values = _daily_token_values(day_ai_messages, day_usage)
        daily_rows.append(
            [
                current,
                len(
                    {
                        customer_id
                        for m in day_messages
                        if (customer_id := conversation_customer_ids.get(m.conversation_id))
                    }
                ),
                len({m.conversation_id for m in day_messages}),
                sum(m.direction == "inbound" for m in day_messages),
                len(day_ai_messages),
                sum(m.sender_type == "manager" for m in day_messages),
                len(day_messages),
                token_coverage.recorded_replies,
                token_coverage.label,
                *token_values,
            ]
        )
    daily_rows.append(
        [
            "ИТОГО ЗА ПЕРИОД",
            len(active_customer_ids),
            len(conversations),
            sum(message.direction == "inbound" for message in messages),
            len(period_ai_messages),
            sum(message.sender_type == "manager" for message in messages),
            len(messages),
            period_coverage.recorded_replies,
            period_coverage.label,
            *_daily_token_values(period_ai_messages, usage_events),
        ]
    )
    _write_sheet(
        daily,
        [
            "Дата",
            "Уникальные клиенты",
            "Активные диалоги",
            "Сообщения клиентов",
            "Ответы Автопилота",
            "Ответы менеджеров",
            "Всего сообщений",
            "AI-ответов с токенами",
            "Статус данных по токенам",
            "Input tokens",
            "Cached input tokens",
            "Cache write tokens",
            "Reasoning tokens",
            "Output tokens (включая reasoning)",
            "Всего токенов",
            "Себестоимость API, ₽",
            "Списание клиенту, ₽",
        ],
        daily_rows,
        date_columns={1},
        currency_columns={16, 17},
        wide_columns={9},
    )
    total_row = daily.max_row
    daily.auto_filter.ref = f"A1:Q{total_row - 1}"
    daily.freeze_panes = "B2"
    for cell in daily[total_row]:
        cell.fill = PatternFill("solid", fgColor="E8F0FE")
        cell.font = Font(bold=True, color="16325C")
        cell.border = Border(top=Side(style="medium", color="8EA9C1"))

    customers_sheet = workbook.create_sheet("Клиенты")
    customer_rows = []

    def customer_name(value: UUID) -> str:
        item = customers.get(value)
        return item.display_name if item else ""

    for customer_id in sorted(active_customer_ids, key=customer_name):
        customer_conversations = [c for c in conversations if c.customer_id == customer_id]
        customer_messages = [
            m for c in customer_conversations for m in by_conversation.get(c.id, [])
        ]
        customer_usage = [
            u for c in customer_conversations for u in usage_by_conversation.get(c.id, [])
        ]
        customer_ai_messages = [m for m in customer_messages if m.sender_type == "ai"]
        customer_coverage = _token_coverage(customer_ai_messages, customer_usage)
        customer_token_values = _daily_token_values(customer_ai_messages, customer_usage)
        customer = customers.get(customer_id)
        customer_channel_types = {
            channel.type
            for c in customer_conversations
            if (channel := channels.get(c.channel_id)) is not None
        }
        customer_rows.append(
            [
                str(customer_id),
                customer.display_name if customer else "",
                identities.get(customer_id, ""),
                ", ".join(sorted(customer_channel_types)),
                len(customer_conversations),
                min((m.created_at for m in customer_messages), default=None),
                max((m.created_at for m in customer_messages), default=None),
                sum(m.direction == "inbound" for m in customer_messages),
                len(customer_ai_messages),
                sum(m.sender_type == "manager" for m in customer_messages),
                len(customer_messages),
                customer_coverage.label,
                *customer_token_values,
            ]
        )
    _write_sheet(
        customers_sheet,
        [
            "Customer ID",
            "Клиент",
            "ID в канале",
            "Каналы",
            "Диалоги",
            "Первое сообщение",
            "Последнее сообщение",
            "Входящие",
            "AI",
            "Менеджеры",
            "Всего сообщений",
            "Статус данных по токенам",
            "Input",
            "Cached",
            "Cache write",
            "Reasoning",
            "Output",
            "Всего токенов",
            "Себестоимость, ₽",
            "Списание, ₽",
        ],
        customer_rows,
        datetime_columns={6, 7},
        currency_columns={19, 20},
        wide_columns={12},
    )

    conversations_sheet = workbook.create_sheet("Диалоги")
    conversation_rows = []
    for conversation in conversations:
        thread = by_conversation.get(conversation.id, [])
        conversation_usage = usage_by_conversation.get(conversation.id, [])
        conversation_ai_messages = [m for m in thread if m.sender_type == "ai"]
        conversation_coverage = _token_coverage(conversation_ai_messages, conversation_usage)
        conversation_token_values = _daily_token_values(
            conversation_ai_messages, conversation_usage
        )
        confidence = [m.confidence for m in thread if isinstance(m.confidence, (int, float))]
        customer = customers.get(conversation.customer_id)
        channel = channels.get(conversation.channel_id)
        conversation_rows.append(
            [
                str(conversation.id),
                str(conversation.customer_id),
                customer.display_name if customer else "",
                channel.type if channel else "",
                conversation.status,
                min((m.created_at for m in thread), default=None),
                max((m.created_at for m in thread), default=None),
                sum(m.direction == "inbound" for m in thread),
                len(conversation_ai_messages),
                sum(m.sender_type == "manager" for m in thread),
                len(thread),
                round(sum(confidence) / len(confidence), 4) if confidence else None,
                ", ".join(sorted({u.model for u in conversation_usage if u.model})),
                conversation_coverage.label,
                *conversation_token_values,
            ]
        )
    _write_sheet(
        conversations_sheet,
        [
            "Conversation ID",
            "Customer ID",
            "Клиент",
            "Канал",
            "Статус",
            "Начало активности",
            "Конец активности",
            "Входящие",
            "AI",
            "Менеджеры",
            "Всего сообщений",
            "Средняя уверенность",
            "Модели",
            "Статус данных по токенам",
            "Input",
            "Cached",
            "Cache write",
            "Reasoning",
            "Output",
            "Всего токенов",
            "Себестоимость, ₽",
            "Списание, ₽",
        ],
        conversation_rows,
        datetime_columns={6, 7},
        percent_columns={12},
        currency_columns={21, 22},
        wide_columns={14},
    )

    if include_messages:
        messages_sheet = workbook.create_sheet("Сообщения")
        usage_by_message = {item.message_id: item for item in usage_events if item.message_id}
        conversation_map = {item.id: item for item in conversations}
        message_rows = []
        for message in messages:
            message_conversation = conversation_map.get(message.conversation_id)
            customer = (
                customers.get(message_conversation.customer_id) if message_conversation else None
            )
            channel = (
                channels.get(message_conversation.channel_id) if message_conversation else None
            )
            usage_event = usage_by_message.get(message.id)
            usage_values = _usage_values_for_message(message, usage_event)
            message_rows.append(
                [
                    message.created_at,
                    str(message.id),
                    str(message.conversation_id),
                    str(message_conversation.customer_id) if message_conversation else "",
                    customer.display_name if customer else "",
                    channel.type if channel else "",
                    message.direction,
                    message.sender_type,
                    message.text,
                    _attachment_names(message.attachments),
                    message.status,
                    message.confidence,
                    str((message.ai_meta or {}).get("decision") or ""),
                    usage_event.provider
                    if usage_event
                    else str((message.ai_meta or {}).get("provider") or ""),
                    usage_event.model
                    if usage_event
                    else str((message.ai_meta or {}).get("model") or ""),
                    *usage_values,
                ]
            )
        _write_sheet(
            messages_sheet,
            [
                "Дата и время",
                "Message ID",
                "Conversation ID",
                "Customer ID",
                "Клиент",
                "Канал",
                "Направление",
                "Отправитель",
                "Сообщение",
                "Вложения",
                "Доставка",
                "Уверенность",
                "Решение",
                "Provider",
                "Модель",
                "Input",
                "Cached",
                "Cache write",
                "Reasoning",
                "Output",
                "Всего токенов",
                "Себестоимость, ₽",
                "Списание, ₽",
            ],
            message_rows,
            datetime_columns={1},
            percent_columns={12},
            currency_columns={22, 23},
            wide_columns={9, 10},
        )

    usage_sheet = workbook.create_sheet("Использование AI")
    usage_rows = [
        [
            u.created_at,
            str(u.id),
            str(u.customer_id or ""),
            str(u.conversation_id or ""),
            str(u.message_id or ""),
            u.provider,
            u.model,
            u.reasoning_effort,
            u.input_tokens,
            u.cached_input_tokens,
            u.cache_write_tokens,
            u.reasoning_tokens,
            u.output_tokens,
            u.total_tokens,
            u.provider_cost_microrubles / 1_000_000,
            u.client_charge_kopecks / 100,
            u.currency_rate_kopecks / 100,
            u.request_id,
        ]
        for u in usage_events
    ]
    _write_sheet(
        usage_sheet,
        [
            "Дата и время",
            "Usage ID",
            "Customer ID",
            "Conversation ID",
            "Message ID",
            "Provider",
            "Модель",
            "Reasoning effort",
            "Input",
            "Cached",
            "Cache write",
            "Reasoning",
            "Output",
            "Всего токенов",
            "Себестоимость, ₽",
            "Списание, ₽",
            "Курс ₽/$",
            "Request ID",
        ],
        usage_rows,
        datetime_columns={1},
        currency_columns={15, 16},
    )
    return workbook


def _group_messages(messages: list[Message]) -> dict[UUID, list[Message]]:
    result: dict[UUID, list[Message]] = {}
    for message in messages:
        result.setdefault(message.conversation_id, []).append(message)
    return result


def _group_usage(events: list[AIUsageEvent]) -> dict[UUID, list[AIUsageEvent]]:
    result: dict[UUID, list[AIUsageEvent]] = {}
    for event in events:
        if event.conversation_id:
            result.setdefault(event.conversation_id, []).append(event)
    return result


def _token_coverage(ai_messages: list[Message], events: list[AIUsageEvent]) -> _TokenCoverage:
    """Describe whether zeroes are measured values or missing historical usage."""
    recorded_message_ids = {event.message_id for event in events if event.message_id is not None}
    recorded_replies = sum(
        message.id in recorded_message_ids
        or _usage_has_tokens((message.ai_meta or {}).get("usage"))
        for message in ai_messages
    )
    if not ai_messages:
        return _TokenCoverage(0, "Нет AI-ответов")
    if recorded_replies == len(ai_messages):
        return _TokenCoverage(recorded_replies, "Полные данные")
    if recorded_replies:
        return _TokenCoverage(
            recorded_replies,
            f"Частично: {recorded_replies} из {len(ai_messages)} AI-ответов",
        )
    return _TokenCoverage(0, "Нет исторических данных")


def _daily_token_values(
    ai_messages: list[Message], events: list[AIUsageEvent]
) -> tuple[
    int | None,
    int | None,
    int | None,
    int | None,
    int | None,
    int | None,
    float | None,
    float | None,
]:
    coverage = _token_coverage(ai_messages, events)
    if coverage.recorded_replies == 0 and ai_messages:
        # Blank cells intentionally mean “not measured”, while numeric zero means
        # the ledger exists and recorded an actual zero for the period/day.
        return (None, None, None, None, None, None, None, None)
    by_message = {event.message_id: event for event in events if event.message_id}
    token_totals = [0, 0, 0, 0, 0, 0]
    for message in ai_messages:
        event = by_message.get(message.id)
        if event:
            values = (
                event.input_tokens,
                event.cached_input_tokens,
                event.cache_write_tokens,
                event.reasoning_tokens,
                event.output_tokens,
                event.total_tokens,
            )
        else:
            usage = (message.ai_meta or {}).get("usage")
            if not _usage_has_tokens(usage):
                continue
            values = (
                _optional_int(usage.get("input_tokens")) or 0,
                _optional_int(usage.get("cached_input_tokens")) or 0,
                _optional_int(usage.get("cache_write_tokens")) or 0,
                _optional_int(usage.get("reasoning_tokens")) or 0,
                _optional_int(usage.get("output_tokens")) or 0,
                _optional_int(usage.get("total_tokens")) or 0,
            )
        token_totals = [total + value for total, value in zip(token_totals, values, strict=True)]
    has_priced_event = bool(events)
    return (
        token_totals[0],
        token_totals[1],
        token_totals[2],
        token_totals[3],
        token_totals[4],
        token_totals[5],
        sum(event.provider_cost_microrubles for event in events) / 1_000_000
        if has_priced_event
        else None,
        sum(event.client_charge_kopecks for event in events) / 100 if has_priced_event else None,
    )


def _response_times(messages: list[Message]) -> list[float]:
    result: list[float] = []
    for thread in _group_messages(messages).values():
        pending: datetime | None = None
        for message in thread:
            if message.direction == "inbound":
                pending = message.created_at
            elif pending and message.sender_type in {"ai", "manager"}:
                result.append(max(0, (message.created_at - pending).total_seconds()))
                pending = None
    return result


def _usage_values_for_message(
    message: Message, event: AIUsageEvent | None
) -> tuple[
    int | None,
    int | None,
    int | None,
    int | None,
    int | None,
    int | None,
    float | None,
    float | None,
]:
    if event:
        return (
            event.input_tokens,
            event.cached_input_tokens,
            event.cache_write_tokens,
            event.reasoning_tokens,
            event.output_tokens,
            event.total_tokens,
            event.provider_cost_microrubles / 1_000_000,
            event.client_charge_kopecks / 100,
        )
    meta_usage = (message.ai_meta or {}).get("usage")
    if not _usage_has_tokens(meta_usage):
        return (None, None, None, None, None, None, None, None)
    return (
        _optional_int(meta_usage.get("input_tokens")),
        _optional_int(meta_usage.get("cached_input_tokens")),
        _optional_int(meta_usage.get("cache_write_tokens")),
        _optional_int(meta_usage.get("reasoning_tokens")),
        _optional_int(meta_usage.get("output_tokens")),
        _optional_int(meta_usage.get("total_tokens")),
        None,
        None,
    )


def _usage_has_tokens(value: object) -> TypeGuard[dict[str, object]]:
    if not isinstance(value, dict):
        return False
    return any(
        _optional_int(value.get(key)) is not None
        for key in (
            "input_tokens",
            "cached_input_tokens",
            "cache_write_tokens",
            "reasoning_tokens",
            "output_tokens",
            "total_tokens",
        )
    )


def _optional_int(value) -> int | None:
    return value if isinstance(value, int) and not isinstance(value, bool) else None


def _attachment_names(attachments: dict) -> str:
    return ", ".join(
        str(item.get("name") or item.get("filename") or "Вложение")
        for item in attachments.get("items", [])
        if isinstance(item, dict)
    )


def _write_sheet(
    sheet,
    headers: list[str],
    rows: list[list],
    *,
    date_columns: set[int] | None = None,
    datetime_columns: set[int] | None = None,
    percent_columns: set[int] | None = None,
    currency_columns: set[int] | None = None,
    wide_columns: set[int] | None = None,
) -> None:
    date_columns = date_columns or set()
    datetime_columns = datetime_columns or set()
    percent_columns = percent_columns or set()
    currency_columns = currency_columns or set()
    wide_columns = wide_columns or set()
    sheet.append(headers)
    for row in rows:
        sheet.append([_excel_value(item) for item in row])
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    sheet.sheet_view.showGridLines = False
    header_fill = PatternFill("solid", fgColor="2463EB")
    header_font = Font(color="FFFFFF", bold=True)
    border = Border(bottom=Side(style="thin", color="D9E1EC"))
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    sheet.row_dimensions[1].height = 42
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=cell.column in wide_columns)
            if cell.column in date_columns:
                cell.number_format = "yyyy-mm-dd"
            elif cell.column in datetime_columns:
                cell.number_format = "yyyy-mm-dd hh:mm:ss"
            elif cell.column in percent_columns:
                cell.number_format = "0.0%"
            elif cell.column in currency_columns:
                cell.number_format = '#,##0.00 "₽"'
    for index, header in enumerate(headers, start=1):
        values = [str(header)] + [
            str(row[index - 1] or "") for row in rows[:200] if index - 1 < len(row)
        ]
        width = min(
            48 if index in wide_columns else 30,
            max(11, max(len(value) for value in values) + 2),
        )
        sheet.column_dimensions[get_column_letter(index)].width = width


def _excel_value(value):
    if isinstance(value, datetime):
        return value.astimezone(ANALYTICS_TIMEZONE).replace(tzinfo=None) if value.tzinfo else value
    return value


def _resolve_period(from_date: date | None, to_date: date | None) -> tuple[date, date]:
    today = datetime.now(ANALYTICS_TIMEZONE).date()
    resolved_to = to_date or today
    resolved_from = from_date or (resolved_to - timedelta(days=DEFAULT_PERIOD_DAYS - 1))

    if resolved_from > resolved_to:
        raise HTTPException(status_code=422, detail="'from' must be before or equal to 'to'")
    if (resolved_to - resolved_from).days >= MAX_PERIOD_DAYS:
        raise HTTPException(
            status_code=422,
            detail=f"Analytics period cannot exceed {MAX_PERIOD_DAYS} days",
        )
    return resolved_from, resolved_to


def _analytics_bounds(date_from: date, date_to: date) -> tuple[datetime, datetime]:
    starts_at = datetime.combine(date_from, time.min, tzinfo=ANALYTICS_TIMEZONE)
    ends_at = datetime.combine(
        date_to + timedelta(days=1),
        time.min,
        tzinfo=ANALYTICS_TIMEZONE,
    )
    return starts_at.astimezone(UTC), ends_at.astimezone(UTC)


def _daily_series(
    inbound_messages: list[Message], date_from: date, date_to: date
) -> list[AnalyticsDailySeriesItem]:
    active_dialogs_by_date: dict[date, set[UUID]] = {}
    for message in inbound_messages:
        message_date = _as_analytics_date(message.created_at)
        active_dialogs_by_date.setdefault(message_date, set()).add(message.conversation_id)
    counts = Counter(
        {
            message_date: len(conversation_ids)
            for message_date, conversation_ids in active_dialogs_by_date.items()
        }
    )
    days_count = (date_to - date_from).days + 1
    return [
        AnalyticsDailySeriesItem(
            date=current_date,
            dialogs=counts[current_date],
        )
        for offset in range(days_count)
        if (current_date := date_from + timedelta(days=offset))
    ]


def _as_analytics_date(value: datetime) -> date:
    if value.tzinfo is None:
        value = value.replace(tzinfo=UTC)
    return value.astimezone(ANALYTICS_TIMEZONE).date()


async def _usage_and_limit(session: SessionDep, tenant_id: UUID) -> tuple[int, int]:
    current_period = datetime.now(UTC).strftime("%Y-%m")
    usage_result = await session.execute(
        select(UsageCounter).where(
            UsageCounter.tenant_id == tenant_id,
            UsageCounter.period == current_period,
        )
    )
    usage = usage_result.scalar_one_or_none()

    subscription_result = await session.execute(
        select(Subscription, Plan)
        .join(Plan, Subscription.plan_id == Plan.id)
        .where(Subscription.tenant_id == tenant_id)
        .order_by(Subscription.created_at.desc())
    )
    subscription_row = subscription_result.first()
    dialogs_limit = subscription_row[1].dialog_limit if subscription_row else 0
    return usage.dialogs_count if usage else 0, dialogs_limit


async def _knowledge_counts(session: SessionDep, tenant_id: UUID) -> tuple[int, int, int]:
    documents_result = await session.execute(
        select(KbDocument).where(
            KbDocument.tenant_id == tenant_id,
            KbDocument.status == "ready",
        )
    )
    chunks_result = await session.execute(select(KbChunk).where(KbChunk.tenant_id == tenant_id))
    candidates_result = await session.execute(
        select(KbCandidate).where(
            KbCandidate.tenant_id == tenant_id,
            KbCandidate.status == "pending",
        )
    )
    return (
        len(documents_result.scalars().all()),
        len(chunks_result.scalars().all()),
        len(candidates_result.scalars().all()),
    )


def _status_counts(conversations: list[Conversation]) -> dict[str, int]:
    counts: dict[str, int] = {}
    for conversation in conversations:
        counts[conversation.status] = counts.get(conversation.status, 0) + 1
    return counts


def _average_response_sec(messages: list[Message]) -> int:
    by_conversation: dict[UUID, list[Message]] = {}
    for message in messages:
        by_conversation.setdefault(message.conversation_id, []).append(message)

    response_times: list[float] = []
    for conversation_messages in by_conversation.values():
        pending_inbound_at: datetime | None = None
        for message in conversation_messages:
            if message.direction == "inbound":
                pending_inbound_at = message.created_at
                continue
            if (
                pending_inbound_at is not None
                and message.direction == "outbound"
                and message.sender_type in {"ai", "manager"}
            ):
                response_times.append(
                    max(0.0, (message.created_at - pending_inbound_at).total_seconds())
                )
                pending_inbound_at = None

    if not response_times:
        return 0
    return round(sum(response_times) / len(response_times))


def _ratio(value: int, total: int) -> float:
    if total <= 0:
        return 0.0
    return round(min(max(value, 0), total) / total, 4)
